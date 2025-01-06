import asyncio
from argparse import ArgumentParser
from playwright.async_api import async_playwright
from pathlib import Path
from openpyxl import load_workbook

SEARCH_CONTINUE_LOCK = asyncio.Event()

SEARCH_KEYBOARDS = []
SEARCH_KEYBOARDS_INDEX = 0

SEARCH_KEYBOARD_MAP = {}


async def click_on_teams_chat_search(page):
    # try to find "Find in chat" menu item directly
    find_in_chat = await page.query_selector(
        'div[role="menuitem"][data-tid="chat-header-find-in-chat-search"]'
    )
    if find_in_chat:
        await find_in_chat.click()
        print("Find in chat menu item found directly")
    else:
        # find "more" button
        more_button = await page.query_selector(
            'button[data-tid="chat-header-more-menu-trigger"]'
        )
        await more_button.click()
        # find "Find in chat" menu item
        find_in_chat = await page.query_selector(
            'div[role="menuitem"][data-tid="chat-header-find-in-chat-search"]'
        )
        await find_in_chat.click()
        print("Find in chat menu item found after clicking more")


async def input_text_to_teams_chat_search(page, text):
    # find chat search input
    chat_search_input = await page.query_selector(
        'input[data-tid="contextual-search-input"]'
    )
    await chat_search_input.fill(text)
    await chat_search_input.press("Enter")


async def on_response(response):
    if response.url != "https://substrate.office.com/search/api/v2/query":
        return
    content = await response.json()
    global SEARCH_KEYBOARDS_INDEX
    keyboard = SEARCH_KEYBOARDS[SEARCH_KEYBOARDS_INDEX]
    word_breaker_language = content["EntitySets"][0]["Properties"][
        "WordBreakerLanguage"
    ]
    SEARCH_KEYBOARD_MAP[keyboard] = word_breaker_language
    SEARCH_KEYBOARDS_INDEX += 1
    SEARCH_CONTINUE_LOCK.set()


def get_search_keyboards(xl_path, sheet_name):
    wb = load_workbook(xl_path)
    ws = wb[sheet_name]
    search_keyboards = []
    for row in ws.iter_rows():
        search_keyboards.append(row[2].value)
    return search_keyboards


def save_test_result(xl_path, sheet_name, search_keyboard_map):
    wb = load_workbook(xl_path)
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        search_keyboard = row[2].value
        word_breaker_language = search_keyboard_map.get(search_keyboard)
        row[1].value = word_breaker_language
    wb.save(xl_path)


async def main():
    parser = ArgumentParser()
    parser.add_argument("xl", type=Path)
    parser.add_argument("wb", type=str)
    parser.add_argument("--save", action="store_true")
    args = parser.parse_args()
    global SEARCH_KEYBOARDS
    SEARCH_KEYBOARDS = get_search_keyboards(args.xl, args.wb)
    print(SEARCH_KEYBOARDS)
    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp("http://localhost:9222")
        default_context = browser.contexts[0]
        page = await default_context.new_page()
        await page.goto("https://teams.microsoft.com/v2/")
        # wait user clict enter to continue
        input("Press Enter to continue...")
        # Monitor requests
        page.on("response", lambda response: asyncio.create_task(on_response(response)))
        await click_on_teams_chat_search(page)
        for keyboard in SEARCH_KEYBOARDS:
            SEARCH_CONTINUE_LOCK.clear()
            await input_text_to_teams_chat_search(page, keyboard)
            print(f"Searching for {keyboard}")
            await SEARCH_CONTINUE_LOCK.wait()

        print(SEARCH_KEYBOARD_MAP)
        # await browser.close()

    if args.save:
        save_test_result(args.xl, args.wb, SEARCH_KEYBOARD_MAP)
        print("Test result saved")


if __name__ == "__main__":
    asyncio.run(main())
