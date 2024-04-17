from openpyxl import load_workbook, Workbook
import argparse
from os.path import expanduser
from pathlib import Path
import re


parser = argparse.ArgumentParser(description="Convert a docx file to an xlsx file")
parser.add_argument("--path", required=True, help="Path to the input file")
parser.add_argument("--input-sheet", default="Sheet1", help="Name of the input sheet")
parser.add_argument("--output", default="output.xlsx", help="Path to the output file")
args = parser.parse_args()
path = expanduser(args.path)
output = expanduser(args.output)
workbook = load_workbook(path)

if not Path(path).exists():
    print("Source file not found")
    exit(1)

worksheet = workbook.get_sheet_by_name(args.input_sheet)

data_map = []

valid_header = ["功能", "子功能"]
valid_header_index = []
for raw in worksheet.iter_rows():
    index = raw.index
    data = []
    for cell in raw:
        cell_value = cell.value
        cell_index = cell.column
        if cell_value in valid_header:
            valid_header_index.append(cell_index)
            continue
        if cell_index in valid_header_index:
            data.append(cell_value)
    data = [i for i in data if i]
    if len(data) == 1:
        data = [data_map[-1][0]] + data
    if data:
        data_map.append(data)

data_map = [[re.sub(r"\(.*\)", "", i) for i in data] for data in data_map if data]
separators = [":", "-", "："]

processed_data_map = []
for data in data_map:
    if data:
        first_item = data[0]
        second_item = data[1]
        for sep in separators:
            if sep in second_item:
                second_item = sep.join(second_item.rsplit(sep, maxsplit=1)[:-1])
                break
        processed_data_map.append([first_item, second_item])
data_map = processed_data_map

r_data_map = []
for data in data_map:
    if data not in r_data_map:
        r_data_map.append(data)

print(r_data_map)

workbook = Workbook()
worksheet = workbook.active
for raw in r_data_map:
    worksheet.append(raw)
merge_range = []
tmp_key = None
i = 0
for data in r_data_map:
    i += 1
    if tmp_key == data[0]:
        merge_range[-1][1] = i
    else:
        tmp_key = data[0]
        merge_range.append([i, i])

print(merge_range)
for m_range in merge_range:
    worksheet.merge_cells(f"A{m_range[0]}:A{m_range[1]}")

workbook.save(output)
